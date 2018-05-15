# -*- coding: utf-8 -*-
"""Scanning a fastq.gz to count DNA barcodes/tag defined in an Excel

This module aims to count occurrence of prefined short DNA sequences
(barcode/tag) and its combinations in a fastq.gz file from Next Generation
Sequencing. 

The input files include a .fastq.gz file and barcode sets in Excel file with
specific format (-b). Other required arguments are library length (insert_size,
-s), direction (-d, 0 - forward, 1 - backward, 2 - both), type of barcode being
put at column in output file (-c), type of barcode being put at row (-r),
number of processes (-p) and number of reads to be scanned (-n). 

The output file is an Excel file with 4 sheets: summary, barcode_count,
barcode_Read-Per-Million, and BarPlot. 

"""

__version__ = '0.1'
__author__ = 'Jingqiao Lu, jingqiao.lv@gmail.com'

import io
import os
import sys
import gzip
import argparse
import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import multiprocessing as mp
from functools import partial

__author__ = 'Jingqiao Lu'

def main():
    """ Entry point of the module for commandline execution.

    Output an excel file after receiving 9 arguments via argparse.

    Args:
        fastq   Fastq.gz file to be scanned, required
        result  output Excel file with suffix .xlsx, required
        barcode_setting Excel file containing barcodes information, 
                        including position, name, type, id and 
                        sequences, required
        insert_size length of library, or DNA to be sequenced,
                    required
        direction    direction of library or DNA to be sequences, 
                    required, 0=forward, 1=barkward, 2=both, default 0
        barcode_at_column   barcode type appears at column in output
                            valid when number of barcode type is 2. 
                            default 'sample'
        barcode_at_row      barcode type appears at row in output file,
                            valid when number of barcode type is 2.
                            default 'target'
        processes   number of processes for multiprocessing, intger, 
                    default 1
        number_reads   number of reads for scanning
   output:
        An Excel file of name specified by argument 'result'.
        The output contains 4 sheets:
            summary:        numer of reads scanned, and reads with barcode
            barcode_count:  count of reads with barcode combinations
            barcode_Read-Per-Million:   normalization of counts within sample
            BarPlot:    barplot of barcode counts and RPM
    """
    parser = argparse.ArgumentParser(
        description='Scanning a gzipped fastq to count barcodes '\
        'provided in an Excel file' )
    parser.add_argument('fastq_gz', help='input gzipped fastq (fastq.gz)')
    parser.add_argument('result', help='result excel file (.xlsx)')
    parser.add_argument('-b', dest='barcode_setting', required=True,
                        help='barcode setting in excel (.xlsx)')
    parser.add_argument('-s', dest='insert_size', default=100, type=int, 
                        help='library insert size, default: 100',
                        )
    parser.add_argument('-d', dest='direction', type=int, default=0,
                        help='library direction: '\
                        '{0 forward 1 backward 2 both}, default: 0')
    parser.add_argument('-c', dest='barcode_at_column', default='sample', 
                        help='barcode type at column in output file, default: '\
                        'sample')
    parser.add_argument('-r', dest='barcode_at_row', default='target', 
                        help='barcode type at row in output file, default: '\
                       'target')
    parser.add_argument('-p', dest='processes', 
                        help='processes number, default: 1', 
                        type=int, default=1)
    parser.add_argument('-n', dest='number_reads', type=int, default=0,
                        help='number of reads for scanning')
    args = parser.parse_args()
    print("\ncountBarcodeInFastq.py -b {} -s {} -d {} -r {} -c {} -p {} -n {}"\
          " {} {}".format(args.barcode_setting, args.insert_size,
                          args.direction, args.barcode_at_row,
                          args.barcode_at_column, args.processes,
                          args.fastq_gz, args.result, args.number_reads ))
    barcode_combination = read_barcode_list_excel(args.barcode_setting,
                                                  args.insert_size)
    print("\nBarcode Type(s): ".format(len(
          barcode_combination.barcode_combination_types())))
    [print(" {:>15}: {:3} - {:3}  {}  {}".format(bs._barcode_set_name,
                                      bs._start_position+1,
                                      bs._end_position,
                                      bs._barcode_type,
                                      len(bs._barcode_sequences)
                                      )) for bs in 
         barcode_combination._barcode_sets]
    print("Barcode Combinations: {}".format(len(
          barcode_combination.barcode_combination_labels()))) 
    result = count_barcode_in_fastq(args.fastq_gz, barcode_combination,
                                    identify_barcode_in_read, args.direction,
                                    args.processes, args.number_reads)
    result = barcode_count_result_rearrange(result, args.barcode_at_column,
                                            args.barcode_at_row)
    print("Reads scanned: {}".format(result["total_reads"]))
    print("Reads with barcodes: {}".format(result["reads_identified"]))
    report_result_in_excel(result, args.result)


def read_barcode_list_excel(barcode_excel, insert_size=100):
    """Read barcode information from Excel file and return a BarcodeCombination

    Args:
        barcode_excel   an Excel file containing barcode information
        insert_size     the length of DNA to be sequenced

    Return:
        A BarcodeCombination object
    """
    barcode_dfs = pd.read_excel(barcode_excel, None)
    df_barcode_setting = barcode_dfs['barcode_position']
    barcode_sets = []
    for idx, row in df_barcode_setting.iterrows():
        barcode_set_name = row['barcode_set']
        if barcode_set_name not in barcode_dfs.keys():
            raise ValueError("No sheet '{}'".format(barcode_set_name))
        bs = None
        bs = BarcodeSet(barcode_set_name, int(row["start"]),
                        int(row["end"]), row["barcode_type"],                                           insert_size)
        for _, barcode_row in barcode_dfs[barcode_set_name].iterrows(): 
            bs.add(barcode_row['barcode_id'], barcode_row['barcode_sequence'])
        barcode_sets.append(bs)
    barcode_sets.sort(key=lambda x: x.start_position())
    return BarcodeCombination(barcode_sets)

def count_barcode_in_fastq(fastq_gz, barcode_combination, idenfunc,
                           direction=0, processes=1, number_reads=0):

    reads, results, barcode_counts = [], [], []
    read_count = 0
    # attach arguments to barcode identification function 
    # (barcode_combination, library direction)
    iden = partial(idenfunc, barcode_combination=barcode_combination,
                   direction=direction)
    def count_barcode_in_reads(reads, processes, iden):
        if processes > 1:
            pool = mp.Pool(processes=processes)
            result = pool.map(iden, reads)
            pool.close()
            pool.join()
        else:
            result = [iden(read) for read in reads]
        return result

    with io.BufferedReader(gzip.open(fastq_gz, 'rb')) as f:
        print("Counting barcodes in "\
              "{}...".format(os.path.basename(fastq_gz))) 
        line_count, read_count = 0, 0
        for line in f:
            line_count +=1
            if line_count % 4 != 2:  # skip non sequence line
                continue
            reads.append(line.decode()[:-1]) # add sequence into reads list
            read_count += 1
            if read_count % 100000 == 0:
                result = count_barcode_in_reads(reads, processes, iden)
                results.extend(result)
                barcode_counts.append(np.apply_along_axis(sum, 0,                                                             np.array(results)))
                reads = []
                results= []
                print('.', end='', flush=True)
            if len(barcode_counts) == 5:
                barcode_counts = [np.apply_along_axis(sum, 0, 
                          np.array(barcode_counts))]
            if number_reads>0 and read_count>=number_reads:
                break
    if reads:
        result = count_barcode_in_reads(reads, processes, iden)
        results.extend(result)
        barcode_counts.append(np.apply_along_axis(sum, 0, np.array(results)))
        print('.', end='', flush=True)
    print()
    barcode_counts = np.apply_along_axis(sum, 0, np.array(barcode_counts))
    return {'total_reads': read_count,
            'barcode_counts': barcode_counts, 
            'barcode_types': barcode_combination.barcode_combination_types(),
            'barcode_labels': barcode_combination.barcode_combination_labels()}

def barcode_count_result_rearrange(barcode_count_result, barcode_at_column,
                                   barcode_at_row):
    df = pd.DataFrame(barcode_count_result['barcode_labels'])
    #barcode_type_at_column, barcode_type_at_row = None, None
    barcode_type_list = []
    for key, val in barcode_count_result['barcode_types'].items():
        barcode_type_list.append(key)
        df[key] = df.iloc[:,val[0]]
        for barcode_order, pos in enumerate(val):
            if len(val) > 1:
                df[key] = df[key] + '_' + df.iloc[:,pos]
  
    df['count'] = barcode_count_result['barcode_counts']
    df =  df[list(barcode_count_result['barcode_types'].keys()) + ['count']]

    df_dimension = 1
    if len(barcode_type_list) == 2:
        df = df.pivot_table(index=barcode_at_row, columns=barcode_at_column,
                            values='count')
        df_dimension = 2

    return {'total_reads': barcode_count_result['total_reads'],
            'barcode_counts': df,
            'reads_identified': barcode_count_result['barcode_counts'].sum(),
            'result_dimension': df_dimension
            }

def report_result_in_excel(rearranged_result, output_file):
    writer = pd.ExcelWriter(output_file)
    df_summary = pd.DataFrame([(rearranged_result['total_reads']),
                               (rearranged_result['reads_identified'])],
                              columns=['count'], index=['total_reads',
                                                      'reads_identified'])
    df_summary.to_excel(writer, 'summary')
    df = rearranged_result['barcode_counts']
    df.to_excel(writer, 'barcodes_counts',
                index=rearranged_result['result_dimension']==2)
    if rearranged_result['result_dimension'] == 1:
        df['RPM'] = df['count'] * 10**6 / df['count'].sum()
        df.to_excel(writer, 'barcodes_Read-Per-Million', index=False)
        two_bar_plots(df['RPM'], df['count'], output_file+'.png')
    else:
        df_rpm = df * 10**6 / df.sum(axis=0)
        df_rpm = df_rpm.fillna(0)
        df_rpm.to_excel(writer, 'barcodes_Read-Per-Million')
        #df_rpm.transpose().plot(kind='bar')#, ylim=(0, df_rpm.max().max()*10))
        two_bar_plots(df.transpose(), df_rpm.transpose(), output_file+'.png')
    writer.save()
    wb = openpyxl.load_workbook(output_file)
    ws = wb.create_sheet('BarPlot')
    img = openpyxl.drawing.image.Image(output_file+'.png')
    img.anchor(ws.cell(row=2, column=2))
    ws.add_image(img)
    wb.save(output_file)
    os.remove(output_file+'.png')
    print("Writing result into '{}'.\n".format(output_file))

def two_bar_plots(count_object, rpm_object, image_name):
    fig, (ax1, ax2) = plt.subplots(2, sharex=True)
    fig.subplots_adjust(hspace=0)
    rpm_object.plot(kind='bar', ax=ax1)
    count_object.plot(kind='bar', ax=ax2).invert_yaxis()
    ax2.set_ylabel("Read Count")
    ax1.set_ylabel("Read Per Million")
    ax1.set_yscale("log", nonposy='clip')
    ax2.set_yscale("log", nonposy='clip')
    ax1.set_title("Barcode Counts Barplot")
    plt.savefig(image_name)


def identify_barcode_in_read(read, barcode_combination, direction=0):
    """Identify occurence of barcode in a DNA sequence (read)
        Args:
            read: A DNA sequence consist of 'A', 'C', 'G', 'T' and 'N'
            barcode_combination: An BarcodeCombination object
        Returns:
            A 0,1 numpy array showing occurrence in read for each
            barcode combination 
    """

    # Get barcode information from BarcodeCombination Object
    barcode_positions = barcode_combination.barcode_combination_positions()
    barcode_combinations = barcode_combination.barcode_combination_in_number() 

    # Retrieve sequence from read at barcode positions
    sequences_at_barcode_positions = [ pattern_to_number(read[pos[0]:pos[1]])
                                      for pos in barcode_positions ]
    # Check occurrence with np.array_qual and np.apply_along_axis
    occurrence = np.apply_along_axis(
          lambda x: np.array_equal(x, 
                                    np.array(sequences_at_barcode_positions)),
                 1, barcode_combinations) + 0
    if direction == 0:  
        return occurrence
    
    # Get reverse barcode information from BarcodeCombination Object
    reverse_barcode_positions = barcode_combination.reverse_barcode_combination_positions()
    reverse_barcode_combinations = barcode_combination.reverse_barcode_combination_in_number()

    # Identify occurrence of barcode in read at reverse direction
    sequences_at_reverse_barcode_positions = [pattern_to_number(
          read[pos[0]:pos[1]]) for pos in
          reverse_barcode_positions  ]
    reverse_occurrence = np.apply_along_axis(
          lambda x: np.array_equal(x, 
                 np.array(sequences_at_reverse_barcode_positions)), 1, 
                 reverse_barcode_combinations) +0
    if direction == 1:
        return reverse_occurrence
    elif direction == 2:
        return occurrence + reverse_occurrence

def combinations(barcode_lists):
    if len(barcode_lists) == 1:
        if isinstance(barcode_lists[0][0], list):
            return barcode_lists[0]
        else:
            return [[x] for x in barcode_lists[0]]
    prefix_list = barcode_lists[0]
    if not isinstance(prefix_list[0], list):
        prefix_list = [[x] for x in barcode_lists[0]]
    suffix_list = barcode_lists[1]
    target_list = []
    for prefix in prefix_list:
        for suffix in suffix_list:
            target_list.append(prefix[:] + [suffix])
    if len(barcode_lists) == 2:
        return target_list
    else:
        return combinations([target_list]+barcode_lists[2:])

def pattern_to_number(pattern):
    symbol_dict = {'A':0, 'C':1, 'G':2, 'T':3, 'N':4}
    if len(pattern) == 1:
        return symbol_dict[pattern]
    return symbol_dict[pattern[0]] * (5**(len(pattern)-1)) + \
                    pattern_to_number(pattern[1:])

def reverse_complement(pattern):
    symbol_dict = {'A':'T', 'C':'G', 'G':'C', 'T':'A', 'N':'N'}
    return ''.join([symbol_dict[base] for base in pattern][::-1])



class DuplicateError(Exception):
    pass
    '''def __init__(self, message, errors):
        super().__init__(message)
        self.errors = errors
    '''


class BarcodeCombination:

    def __init__(self, barcode_sets):
        self._barcode_combination_types = [x.barcode_type() for x in
                                           barcode_sets]
        self._barcode_combination_types_dict = self.list2dict(
              self._barcode_combination_types)
        self._barcode_combination_labels = combinations( 
              [x.barcode_labels() for x in barcode_sets])
        self._barcode_combination_sequences = combinations(
              [x.barcode_sequences() for x in barcode_sets])
        self._barcode_combination_in_number = np.array(combinations(
              [x.barcode_in_number() for x in barcode_sets]))
        self._barcode_combination_positions = [(x.start_position(),
                                                x.end_position()) for x in
                                               barcode_sets]
        self._reverse_barcode_combination_positions =[
              (x.reverse_start_position(), x.reverse_end_position()) for x in
              barcode_sets]
        self._reverse_barcode_combination_in_number = np.array(combinations(
              [x.reverse_barcode_in_number() for x in barcode_sets ]))
        self._barcode_sets = barcode_sets

    def list2dict(self, l):
        my_dict = {}
        for i, item in enumerate(l):
            my_dict[item] = my_dict.get(item, []) + [i]
        return my_dict

    def barcode_combination_positions(self):
        return self._barcode_combination_positions
    
    def reverse_barcode_combination_positions(self):
        return self._reverse_barcode_combination_positions

    def barcode_combination_sequences(self):
        return self._barcode_combination_sequences

    def barcode_combination_in_number(self):
        return self._barcode_combination_in_number

    def reverse_barcode_combination_in_number(self):
        return self._reverse_barcode_combination_in_number

    def barcode_combination_labels(self):
        return self._barcode_combination_labels

    def barcode_combination_types(self):
        return self._barcode_combination_types_dict



class BarcodeSet:
    """Return a barcode Set 
       
       Args:
           barcode_set_name: string
           start_position: start position of barcode in library
           end_position: end position of barcode in library
           insert_size: length of library to be sequenced
      
       Return:
           A barcode set with multiple attributes:
               barcode_set_name
               start_position
               end_position
               barcode_labels: A list
               barcode_sequences: a list
               barcode_in_number: conversion barcode sequences into integers
               reverse_start_position
               reverse_end_position
               reverse_barcode_in_number
    """

    def __init__(self, barcode_set_name,start_position, end_position,
                 barcode_type, insert_size):
        self._barcode_set_name = barcode_set_name
        self._barcode_type = barcode_type
        self._insert_size = insert_size
        self._start_position = start_position - 1
        self._end_position = end_position
        self._barcode_length = self._end_position - self._start_position
        self._reverse_start_position = insert_size - end_position
        self._reverse_end_position = insert_size - (start_position-1)
        self._barcode_sequences = []
        self._barcode_labels = []
        self._barcode_in_number = []
        self._reverse_barcode_in_number = []
        self._barcode_number = 0
       
    def add(self, barcode_label, barcode_sequence):
        if barcode_label in self._barcode_labels:
            raise DuplicateError("Duplicate barcode label '{}'!".
                              format(barcode_label))
        if barcode_sequence in self._barcode_sequences:
            raise DuplicateError("Duplicate barcode sequence '{}'!".
                              format(barcode_sequence))
        if len(barcode_sequence) != self._barcode_length:
            raise ValueError("Length of Barcode '{0}' is not consistent with"\
                             " barcode setting {1}".format(barcode_sequence,
                                                          self._barcode_length))

        self._barcode_sequences.append(barcode_sequence)
        self._barcode_labels.append(barcode_label)
        self._barcode_in_number.append(pattern_to_number(barcode_sequence))
        self._reverse_barcode_in_number.append(
            pattern_to_number(reverse_complement(barcode_sequence)))
        self._barcode_number += 1

    def remove(self, query):
        index_to_be_removed = None
        try:
            index_to_be_removed = self._barcode_labels.index(query)
        except ValueError:
            pass
        try:
            index_to_be_removed = self._barcode_sequences.index(query)
        except ValueError:
            pass
        if index_to_be_removed:
            del self._barcode_labels[index_to_be_removed]
            del self._barcode_sequences[index_to_be_removed]
            del self._barcode_in_number[index_to_be_removed]
            del self._reverse_barcode_in_number[index_to_be_removed]
            return True
        else:
            print("barcode '{}' can not be found in list".format(query))
            return

    def barcode_labels(self):
        return self._barcode_labels

    def barcode_sequences(self):
        return self._barcode_sequences

    def barcode_in_number(self):
        return self._barcode_in_number

    def reverse_barcode_in_number(self):
        return self._reverse_barcode_in_number

    def start_position(self):
        return self._start_position

    def end_position(self):
        return self._end_position

    def reverse_start_position(self):
        return self._reverse_start_position

    def reverse_end_position(self):
        return self._reverse_end_position

    def barcode_number(self):
        return self._barcode_number

    def barcode_set_name(self):
        return self._barcode_set_name

    def insert_size(self):
        return self._insert_size

    def barcode_type(self):
        return self._barcode_type

if __name__ == '__main__':
    main()
